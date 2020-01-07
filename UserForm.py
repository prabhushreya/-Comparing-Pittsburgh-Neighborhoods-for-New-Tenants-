################## UserForm.py #######################
# Takes in user input through a form and stores it in excel file
######### Group 3, Section C3 ################
#Members:
# Malvika Singh
# Maxwell Kennady
# Shikha Goel
# Sean Hoover
# Poojitha Prasad
# Shreya Prabhu

# import openpyxl and tkinter modules
from openpyxl import *
from tkinter import *
from tkinter import messagebox
import tkinter as tk
# globally declare wb and sheet variable

# opening the existing excel file
wb = load_workbook('C:\\Users\\Public\\UserInput.xlsx') #change path according to your machine

# create the sheet object
sheet = wb.active

def excel():

	# resize the width of columns in
	# excel spreadsheet
	sheet.column_dimensions['A'].width = 30
	sheet.column_dimensions['B'].width = 10
	sheet.column_dimensions['C'].width = 10
	sheet.column_dimensions['D'].width = 20
	sheet.column_dimensions['E'].width = 20
	sheet.column_dimensions['F'].width = 40
	sheet.column_dimensions['G'].width = 50

	# write given data to an excel spreadsheet
	# at particular location
	sheet.cell(row=1, column=1).value = "Neighborhood"
	sheet.cell(row=1, column=2).value = "Monthly Rent($)"
	

# Function to set focus (cursor)
def focus1(event):
	# set focus on the rent_field box
	rent0_field.focus_set()


# Function to set focus
def focus2(event):
	# set focus on the sem_field box
	sem_field.focus_set()


# Function to set focus
def focus3(event):
	# set focus on the form_no_field box
	form_no_field.focus_set()


# Function to set focus
def focus4(event):
	# set focus on the contact_no_field box
	contact_no_field.focus_set()


# Function to set focus
def focus5(event):
	# set focus on the email_id_field box
	email_id_field.focus_set()


# Function to set focus
def focus6(event):
	# set focus on the address_field box
	address_field.focus_set()


# Function for clearing the
# contents of text entry boxes
def clear():

	# clear the content of text entry box
	name_field.delete(0, END)
	rent_field.delete(0, END)
	sem_field.delete(0, END)
	


# Function to take data from GUI where user input cannot be blank or a non-integer field
# window and write to an excel file
def insert():
        if (name_field.get() == "" and rent_field.get() == ""):
                messagebox.showerror("Error", "Please input rent")
        elif (not (rent_field.get().isnumeric())):
                # message box display
                messagebox.showerror("Error", "Please enter a valid rent value")
        else:
                if((int(rent_field.get())>2000) or (int(rent_field.get())<300)):
                        messagebox.showerror("Error", "Rent is not in a valid range [$300-$2000]")
                else:
                        current_row = sheet.max_row
                        current_column = sheet.max_column
                        sheet.cell(row=current_row + 1, column=1).value = popupMenu.selection_get()
                        sheet.cell(row=current_row + 1, column=2).value = rent_field.get()
                        wb.save('C:\\Users\\Public\\UserInput.xlsx')
		


# Driver code
if __name__ == "__main__":
    
    # create a GUI window
    root = Tk()

    # set the background colour of GUI window
    root.configure(background='light blue')

    # set the title of GUI window
    root.title("registration form")

    # set the configuration of GUI window
    root.geometry("500x300")

    excel()

    # create a Form label
    heading = Label(root, text="Welcome to L-ease!",font='Helvetica 16 bold', bg="light blue")
    
    # create a rent label
    rent = Label(root,text="Monthly Rent($)",bg="light blue")


    # create a Name label
    name = Label(root, text="Neighborhood",bg="light blue").grid(row = 2, column = 0)
    tkvar = StringVar(root)
 
    # Dictionary with options
    choices = { 'Shadyside:','North Oakland:','South Oakland:','West Oakland:','Squirrel Hill North:','Squirrel Hill South:','Allegheny Center:','Allentown:','Bloomfield:',
                'Arlington:','South Side Flats:','Mount Washington:','Morningside:','Middle Hill:','Marshall-Shadeland','Manchester:','Lower Lawrenceville:',
                'Lincoln-Lemington-Belmar:','Lincoln Place:','Larimer:','Knoxville:','Hazelwood:','Homewood North:','Overbrook:','Point Breeze:','Stanton Heights:',
                'Perry South:','Greenfield:','Glen Hazel:','Garfield:','Friendship:','Fineview:','Fairywood:','Esplen:','Elliott:','East Liberty:',
                'East Hills:','East Carnegie:','East Allegheny:','Duquesne Heights:','Crawford-Roberts:','Crafton Height:','Chateau:','Chartiers City:',
                'Central Oakland:','Central Northside:','Central Lawrenceville:','Central Business District:','Brookline:','Beechview:','Windgap:','Highland Park:'}
    tkvar.set('ShadySide') 
     
    popupMenu = Listbox(root, selectmode='multiple')
    popupMenu.insert('end', *choices)
    popupMenu.grid(row = 2, column =1)
     
    # on change dropdown value
    def change_dropdown(*args):
        print( tkvar.get() )
     
    # link function to change dropdown
    tkvar.trace('w', change_dropdown)
    
    
    heading.grid(row=0, column=1)
    #name.grid(row=1,column=0)
    rent.grid(row=4, column=0)

    # create a text entry box
    # for typing the information
    name_field = Entry(root)
    rent_field = Entry(root)


    # bind method of widget is used for
    # the binding the function with the events

    # whenever the enter key is pressed
    # then call the focus1 function
    name_field.bind("<Return>", focus1)

    # whenever the enter key is pressed
    # then call the focus2 function
    rent_field.bind("<Return>", focus2)

    
    rent_field.grid(row=4, column=1, ipadx="100")
   
    # call excel function
    excel()

    # create a Submit Button and place into the root window
    submit = Button(root, text="Submit", fg="Black",
                            bg="Red", command=insert)
    submit.grid(row=9, column=1)

    # start the GUI
    root.mainloop()
